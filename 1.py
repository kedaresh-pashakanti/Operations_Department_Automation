import re
from io import BytesIO, StringIO
from decimal import Decimal

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# =============================================================================
# App configuration
# =============================================================================
st.set_page_config(
    page_title="Statement Processor",
    page_icon="📄",
    layout="wide",
)

st.title("📄 Statement Processor")
st.caption("Step 1: Process raw files → Step 2: Append into target Excel file")


# =============================================================================
# Final output structure required by you
# =============================================================================
TARGET_COLUMNS = [
    "Transaction Date",
    "Description",
    "Amount",
    "C.D.Falg",
    "Reference No",
    "Value Date",
    "Branch Name",
    "Running Balance",
]


# =============================================================================
# Helper functions
# =============================================================================
def normalize_col_name(col):
    """Normalize a column/header name for easy matching."""
    return re.sub(r"[^a-z0-9]+", "", str(col).strip().lower())


def clean_reference_no(value):
    """
    Keep Reference No exactly as text so Excel does not truncate it.
    Removes only useless trailing .0 from numeric-looking values.
    """
    if value is None or pd.isna(value):
        return ""

    s = str(value).strip()
    if s.lower() in {"nan", "none", ""}:
        return ""

    # Convert 123456.0 -> 123456 only if it is purely numeric
    if s.endswith(".0"):
        core = s[:-2]
        if core.replace("-", "").isdigit():
            return core

    return s


def find_csv_header_row(raw_bytes, max_lines=120):
    """
    Detect the row number where actual headers start in a CSV file.
    This is useful when the file has summary lines above the table.
    """
    text = raw_bytes.decode("utf-8-sig", errors="ignore")
    lines = text.splitlines()[:max_lines]

    for idx, line in enumerate(lines):
        norm = normalize_col_name(line)

        # This row should contain the important header words
        score = 0
        for token in [
            "transactiondate",
            "transactiondescription",
            "transactionamount",
            "debitcredit",
            "referenceno",
            "valuedate",
            "transactionbranch",
            "runningbalance",
        ]:
            if token in norm:
                score += 1

        if score >= 5:
            return idx

    return None


def find_xlsx_header_row(raw_bytes, max_rows=120):
    """
    Detect the row number where actual headers start in an Excel file.
    """
    wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1
    ):
        values = [normalize_col_name(v) for v in row if v is not None]

        score = 0
        for token in [
            "transactiondate",
            "transactiondescription",
            "transactionamount",
            "debitcredit",
            "referenceno",
            "valuedate",
            "transactionbranch",
            "runningbalance",
        ]:
            if any(token in v for v in values):
                score += 1

        if score >= 5:
            return row_idx

    return None


def read_raw_statement(uploaded_file):
    """
    Read raw CSV / XLSX statement file and detect the real table header row.
    """
    filename = uploaded_file.name.lower()
    raw_bytes = uploaded_file.getvalue()

    if filename.endswith(".csv"):
        header_row = find_csv_header_row(raw_bytes)
        if header_row is None:
            raise ValueError("Header row not found in CSV file.")

        text = raw_bytes.decode("utf-8-sig", errors="ignore")
        df = pd.read_csv(
            StringIO(text),
            skiprows=header_row,
            dtype=str,
            keep_default_na=False,
            na_filter=False,
        )

    elif filename.endswith((".xlsx", ".xlsm")):
        header_row = find_xlsx_header_row(raw_bytes)
        if header_row is None:
            raise ValueError("Header row not found in Excel file.")

        df = pd.read_excel(
            BytesIO(raw_bytes),
            skiprows=header_row - 1,
            dtype=str,
            engine="openpyxl",
            keep_default_na=False,
        )

    else:
        raise ValueError("Only CSV, XLSX, or XLSM files are supported.")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def standardize_statement_df(df):
    """
    Rename raw file columns into the final structure required by the project.
    """
    # Map raw headers to final headers
    rename_map = {
        "Transaction Date": "Transaction Date",
        "Transaction Description": "Description",
        "Transaction Amount": "Amount",
        "Debit / Credit": "C.D.Falg",
        "Reference No.": "Reference No",
        "Reference No": "Reference No",
        "Value Date": "Value Date",
        "Transaction Branch": "Branch Name",
        "Branch Name": "Branch Name",
        "Running Balance": "Running Balance",
    }

    normalized_map = {
        normalize_col_name(k): v for k, v in rename_map.items()
    }

    out = pd.DataFrame()

    for col in df.columns:
        key = normalize_col_name(col)
        if key in normalized_map:
            out[normalized_map[key]] = df[col].astype(str)

    # Add missing columns so output always has the same structure
    for col in TARGET_COLUMNS:
        if col not in out.columns:
            out[col] = ""

    out = out[TARGET_COLUMNS].copy()

    # Keep Reference No as text and avoid truncation
    out["Reference No"] = out["Reference No"].apply(clean_reference_no)

    # Remove fully empty rows
    out = out.replace(r"^\s*$", "", regex=True)
    out = out.loc[~(out == "").all(axis=1)].reset_index(drop=True)

    return out


def find_last_data_row(ws):
    """
    Find the last row that actually contains data.
    This is safer than only relying on ws.max_row.
    """
    for row_idx in range(ws.max_row, 0, -1):
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in row_values):
            return row_idx
    return 0


def append_to_workbook(target_bytes, append_df, sheet_name=None):
    """
    Append standardized data into the selected workbook sheet.
    Returns updated workbook bytes.
    """
    wb = load_workbook(BytesIO(target_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    last_row = find_last_data_row(ws)

    # If sheet is empty, create the headers first
    if last_row == 0:
        ws.append(TARGET_COLUMNS)
        last_row = 1

    # Append new rows with precision handling
    for row in append_df[TARGET_COLUMNS].itertuples(index=False, name=None):
        next_row = ws.max_row + 1

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=next_row, column=col_idx)
            col_name = TARGET_COLUMNS[col_idx - 1]
            col_name_norm = normalize_col_name(col_name)

            # 🔥 FIX: handle Amount / Transaction Amount with precision
            if "AMOUNT" in col_name_norm:
                try:
                    cell.value = Decimal(str(value))
                    # cell.value = float(Decimal(str(value)))
                except:
                    cell.value = value
                cell.number_format = '#,##0.00'
            else:
                cell.value = value

            # Force Reference No column to remain text
            if col_name == "Reference No":
                cell.number_format = "@"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_sheet_names_from_xlsx(uploaded_file):
    """Read sheet names from uploaded target workbook."""
    raw_bytes = uploaded_file.getvalue()
    wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    return wb.sheetnames


# =============================================================================
# Session state
# =============================================================================
if "processed_df" not in st.session_state:
    st.session_state.processed_df = None


# =============================================================================
# Step 1: Raw file processing
# =============================================================================
st.subheader("Step 1 — Process raw files")

raw_files = st.file_uploader(
    "Upload raw statement file(s)",
    type=["csv", "xlsx", "xlsm"],
    accept_multiple_files=True,
    help="Upload one or more raw files. The app will detect the real header row automatically.",
)

#process_clicked = st.button("Process Raw Files", type="primary")
process_clicked = st.button("Process Raw Files")

if process_clicked:
    if not raw_files:
        st.error("Please upload at least one raw file first.")
    else:
        processed_parts = []
        file_errors = []

        with st.spinner("Processing raw file(s)..."):
            for file in raw_files:
                try:
                    raw_df = read_raw_statement(file)
                    std_df = standardize_statement_df(raw_df)
                    processed_parts.append(std_df)
                except Exception as e:
                    file_errors.append(f"{file.name}: {e}")

        if file_errors:
            for msg in file_errors:
                st.error(msg)

        if processed_parts:
            combined_df = pd.concat(processed_parts, ignore_index=True)
            st.session_state.processed_df = combined_df

            st.success(f"Processed successfully. Total rows: {len(combined_df)}")
            st.dataframe(combined_df)
            #st.dataframe(combined_df, use_container_width=True)

            # Optional preview download of processed data
            preview_buf = BytesIO()
            with pd.ExcelWriter(preview_buf, engine="openpyxl") as writer:
                combined_df.to_excel(writer, index=False, sheet_name="Processed_Data")
            preview_buf.seek(0)

            st.download_button(
                "Download Processed Data",
                data=preview_buf.getvalue(),
                file_name="processed_statement.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# =============================================================================
# Step 2: Append to target workbook
# =============================================================================
#st.divider()
st.markdown("---")
st.subheader("Step 2 — Append processed data into target Excel file")

target_file = st.file_uploader(
    "Upload target Excel file",
    type=["xlsx", "xlsm"],
    key="target_file",
    help="This is the file where processed data will be appended.",
)

sheet_name = None
if target_file is not None:
    try:
        sheet_names = get_sheet_names_from_xlsx(target_file)
        sheet_name = st.selectbox("Select target sheet", sheet_names, index=0)
    except Exception as e:
        st.error(f"Unable to read workbook sheets: {e}")

append_clicked = st.button("Append and Create Download")

if append_clicked:
    if st.session_state.processed_df is None:
        st.error("First process the raw file(s) in Step 1.")
    elif target_file is None:
        st.error("Please upload the target Excel file.")
    else:
        try:
            with st.spinner("Appending data to target workbook..."):
                target_bytes = target_file.getvalue()
                updated_bytes = append_to_workbook(
                    target_bytes=target_bytes,
                    append_df=st.session_state.processed_df,
                    sheet_name=sheet_name,
                )

            st.success("Data appended successfully.")

            output_name = f"updated_{target_file.name}"
            st.download_button(
                "Download Updated Excel File",
                data=updated_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Failed to append data: {e}")
