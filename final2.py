"""
HDFC ESCROW MID MAPPING processor

Changes in this version:
- Supports template upload in Streamlit UI
- Shows only rows with blank Transaction Tag in the UI preview
- Applies rule gating by C.D.Falg:
  * MPR Credit From SP -> only when C.D.Falg = C
  * Chargeback / Payout / MDR / Refund -> only when C.D.Falg = D
  * FD -> only when C.D.Falg = C
- Knock Off overrides existing tags
- Keeps other sheets/formulas intact
"""

import os
import sys
import re
import io
import tempfile
from decimal import Decimal, InvalidOperation
from collections import defaultdict

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------
# User-editable defaults for Spyder / CLI
# ---------------------------------------------------------------------
TEMPLATE_PATH = r"Z:\OPERATIONS\Kedaresh\1\main file\HDFC ESCROW MID MAPPING Blank.xlsx"
INPUT_FILE_PATH = r""
OUTPUT_FILE_PATH = r""

TARGET_SHEET_NAME = "ST - HDFC ESCROW"
RAW_COL_COUNT = 8

DEST_TAG_COL = 9
DEST_SP_COL = 10
DEST_SPLIT_REFUND_COL = 11

# ---------------------------------------------------------------------
# Core MPR Credit From SP rules
# The order matters: the first match wins.
# ---------------------------------------------------------------------
MPR_RULES = [
    (["CR-IDIB000F523-ONEPAY MOBILEWARE PRIVATE LIMITED"], "1-INDIANBANKUPI"),
    (["TERMINAL 1 CARDS SETTL."], "2-HDFC"),
    (["RTGS CR-SBIN0004292-STATE BANK OF INDIA"], "3-SBI Acquiring"),
    (["NEFT CR-SBIN0016209-STATE BANK OF INDIA"], "4-SBI NB"),
    (["NEFT CR-ICIC0000018-NDPS"], "5-Atom NB"),
    (["SETTLEMENT ROU"], "6-HDFC NB"),
    (["RTGS CR-UTIB0000100-1PAY MOBILEWARE PVT LTD ESCROW"], "7-AXIS Bank NB"),
    (["CR-YESB0000402-1PAY MOBILEWARE PVT LTD ESCROW"], "8-YES Bank NB"),
    (["CENTRAL LIABILITY OPERATIONS CPC CL"], "9-ICICI NB"),
    (["UPI SETTLEMENT"], "10-HDFCUPI"),
    (["TO CHECK CREDIT NARRATION AND MAP"], "11-ECMS"),
    (["NEFT CR-ICIC0000105-WORLDLINE EPAYMENTS INDIA PRIVATE LIMITED"], "12-WorldLine NB"),
    ([
        "ICIC0099999-ICICI BANK DISB ACC INHOUSE MER ACQ",
        "ICIC0099999-WORLDLINE EPAYMENTS",
    ], "13-ICICICards"),
    (["TO CHECK CREDIT NARRATION AND MAP"], "14-PayzApp"),
    (["1PAYM"], "15-1PayecmsHDFC"),
    (["NO CREDIT IN HDFC ESCROW"], "16-1PayecmsIndianbank"),
    (["NO CREDIT IN HDFC ESCROW"], "17-Airtelpay"),
    ([
        "NEFT CR-CITI0100000-INDIAIDEAS.COM LIMITED",
        "CITI0100000-INDIAIDEAS.COM",
    ], "18-Billdesk"),
    (["NO CREDIT IN HDFC ESCROW"], "19-MobilewareUPI"),
    ([
        "KKBK0000958-NEFT POS ACQUIRING RECEIVABLES",
        "UPI MERCHANT ACQUIRING RECEIVABL",
    ], "20-KotakUPI"),
]

# ---------------------------------------------------------------------
# Extra SP Identifier/MID mapping rules
# ---------------------------------------------------------------------

EXTRA_SP_MID_RULES = [
    ("76034657", "M00028"),
    ("76045442", "M00066"),
    ("70036473", "M000125"),
    ("70036474", "M000123"),
    ("70036475", "M000124"),
    ("76027802", "M00015"),
    ("70044094", "M00006173"),
    ("70039039", "M00005451")
]


def build_refund_rrn_map(refund_df):
    """
    Build a lookup of HDFC UPI refund RRN -> External TID.

    Expected columns in the uploaded HDFC UPI refund file:
    - Txn ref no. (RRN)
    - External TID

    Column matching is done case-insensitively.
    """
    mapping = {}

    if refund_df is None or refund_df.empty:
        return mapping

    rrn_col = None
    external_tid_col = None

    for col in refund_df.columns:
        col_norm = normalize_text(col)
        if col_norm == "TXN REF NO. (RRN)":
            rrn_col = col
        elif col_norm == "EXTERNAL TID":
            external_tid_col = col

    if not rrn_col or not external_tid_col:
        return mapping

    for _, row in refund_df.iterrows():
        rrn = safe_str(row.get(rrn_col, "")).strip()
        external_tid = safe_str(row.get(external_tid_col, "")).strip()
        if rrn and external_tid:
            mapping[rrn] = external_tid

    return mapping


def build_refund_rrn_map_from_paths(refund_paths):
    """Merge multiple uploaded HDFC UPI refund files into one RRN -> External TID map."""
    merged = {}
    if not refund_paths:
        return merged
    for path in refund_paths:
        if not path or not os.path.exists(path):
            continue
        refund_df = read_input_file(path)
        merged.update(build_refund_rrn_map(refund_df))
    return merged


def resolve_refund_sp_from_rrn(reference_no, refund_rrn_map):
    """
    Resolve the SP Identifier/MID for Refund rows by:
    1) matching Reference No against Txn ref no. (RRN)
    2) reading External TID from the refund file
    3) mapping that External TID against EXTRA_SP_MID_RULES
    """
    rrn = safe_str(reference_no).strip()
    
    try:
        rrn = str(int(float(rrn)))
        
    except:
        pass
            
        
    if not rrn or not refund_rrn_map:
        return None

        # 🔥 STEP 2: try direct match 
    external_tid = refund_rrn_map.get(rrn)
    
    # 🔥 STEP 3: fallback (handle mismatch cases)

    
    if not external_tid:
        for key in refund_rrn_map:
            if rrn in key or key in rrn:
                external_tid = refund_rrn_map[key]
                break
            
            if not external_tid:
                return None


    external_tid_norm = normalize_text(external_tid)
    
    # 🔥 STEP 4: mapping

    for needle, sp_mid in EXTRA_SP_MID_RULES:
        if needle in external_tid_norm:
            return sp_mid

    return None


# ---------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------
def normalize_text(value):
    """Uppercase, trim, and collapse whitespace for safe substring checks."""
    if value is None:
        return ""
    text = str(value).replace("\u00A0", " ")
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def parse_amount(value):
    """
    Convert different amount formats into Decimal for safe comparisons.
    Examples:
      2,076.80 -> Decimal('2076.80')
      (2,076.80) -> Decimal('-2076.80')
    """
    if value is None or value == "":
        return None

    text = str(value).strip()
    text = text.replace(",", "")
    text = text.replace("₹", "")
    text = text.replace("INR", "")
    text = text.replace(" ", "")

    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if match:
            try:
                return Decimal(match.group(0))
            except (InvalidOperation, ValueError):
                return None
    return None


def safe_str(value):
    return "" if value is None else str(value)


def extract_m_identifier(description_text):
    match = re.search(r"\bM\d+\b", description_text)
    return match.group(0) if match else ""


def row_tag_allowed(flag, rule_tag):
    """
    Enforce the requested gate:
    - MPR only when C.D.Falg = C
    - Chargeback / Payout / MDR / Refund only when C.D.Falg = D
    - FD only when C.D.Falg = C
    """
    f = normalize_text(flag)

    if rule_tag == "MPR":
        return f == "C"
    if rule_tag in {"CHARGEBACK", "PAYOUT", "MDR", "REFUND"}:
        return f == "D"
    if rule_tag == "FD":
        return f == "C"
    return False


# ---------------------------------------------------------------------
# Rule functions
# ---------------------------------------------------------------------
def get_mpr_credit_from_sp(description):
    text = normalize_text(description)

    for patterns, sp_id in MPR_RULES:
        for pattern in patterns:
            if pattern in text:
                return "MPR Credit From SP", sp_id

    return None, None


def get_chargeback(description):
    text = normalize_text(description)

    if "76027802" in text:
        return "Chargeback", "M00015"

    if "CHARGEBACK" in text:
        return "Chargeback", ""

    return None, None


def get_payout(description):
    text = normalize_text(description)

    if text.startswith("NEFT") or text.startswith("FT"):
        return "Payout", extract_m_identifier(text)

    return None, None


def get_mdr(description):
    text = normalize_text(description)

    if text.startswith("FT") and (
        "MDR-PG" in text
        or "MDR-ERP" in text
        or re.search(r"\bMDR\b", text)
    ):
        return "MDR", "MDR"

    return None, None


def get_fd_mapping(description):
    text = normalize_text(description)

    if (
        "ESCROW TD REDEMPTION PRINCIPAL" in text
        or "ESCROW TD REDEMPTION INTEREST" in text
    ):
        return "FD", "FD"

    return None, None


# def get_refund(description, current_tag=""):
#     """
#     Refund rules:
#     - Description starts with UPI- or UPIREF
#     - Description contains Cr.Voucher Processed
#     - Description ends with Order Refund / Orde Refund
#     - "1CREDIT VOUCHER" only when current tag is empty
#     """
#     text = normalize_text(description)
#     current_tag = (current_tag or "").strip()

#     if text.startswith("UPI-") or text.startswith("UPIREF"):
#         return "Refund", ""

#     if "CR.VOUCHER PROCESSED" in text:
#         return "Refund", ""

#     if re.search(r"\b(ORDER|ORDE)\s+REFUND\s*$", text):
#         return "Refund", ""

#     if not current_tag and "1CREDIT VOUCHER" in text:
#         return "Refund", ""

#     return None, None
def get_refund(description, current_tag="", current_sp=""):
    text = normalize_text(description)
    current_tag = (current_tag or "").strip()
    current_sp = (current_sp or "").strip()

    if text.startswith("UPI-") or text.startswith("UPIREF"):
        return "Refund", ""

    if "CR.VOUCHER PROCESSED" in text:
        return "Refund", ""

    if re.search(r"\b(ORDER|ORDE)\s+REFUND\s*$", text):
        return "Refund", ""

    # 🔥 IMPORTANT FIX
    if "1CREDIT VOUCHER" in text:
        return "Refund", ""   # always override

    return None, None








def get_sp_identifier_mid_mapping(description):
    text = normalize_text(description)

    for needle, sp_mid in EXTRA_SP_MID_RULES:
        if needle in text:
            return sp_mid

    return None


def mark_knock_off_rows(df):
    """
    Knock Off:
      1) same Reference No twice
      2) one C and one D
      3) same amount

    Overrides existing tags, because Knock Off should win.
    """
    needed = {"Reference No", "C.D.Falg", "Amount"}
    if not needed.issubset(df.columns):
        return df

    groups = defaultdict(list)
    for idx, ref in df["Reference No"].items():
        ref_key = safe_str(ref).strip()
        if ref_key:
            groups[ref_key].append(idx)

    for _, row_indices in groups.items():
        if len(row_indices) != 2:
            continue

        i1, i2 = row_indices
        c1 = normalize_text(df.at[i1, "C.D.Falg"])
        c2 = normalize_text(df.at[i2, "C.D.Falg"])
        if {c1, c2} != {"C", "D"}:
            continue

        amt1 = parse_amount(df.at[i1, "Amount"])
        amt2 = parse_amount(df.at[i2, "Amount"])
        if amt1 is None or amt2 is None or amt1 != amt2:
            continue

        for i in (i1, i2):
            df.at[i, "Tranaction Tag"] = "Knock Off"
            df.at[i, "SP Identifier/MID"] = "Knock Off"

    return df


# ---------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------
def read_input_file(input_file_path):
    ext = os.path.splitext(input_file_path)[1].lower()

    if ext in [".xlsx", ".xlsm", ".xls"]:
        wb = openpyxl.load_workbook(input_file_path, data_only=False)
        ws = wb[wb.sheetnames[0]]

        data = list(ws.iter_rows(values_only=True))
        if not data:
            return None

        header = list(data[0])
        rows = data[1:]
        if len(header) == 0:
            return None

        columns = []
        for i, h in enumerate(header):
            if h is None or str(h).strip() == "":
                columns.append(f"col_{i+1}")
            else:
                columns.append(str(h).strip())

        out = {col: [] for col in columns}
        for row in rows:
            row = list(row)
            if len(row) < len(columns):
                row.extend([None] * (len(columns) - len(row)))
            for i, col in enumerate(columns):
                out[col].append(row[i])

        return pd.DataFrame(out)

    if ext == ".csv":
        return pd.read_csv(input_file_path)

    raise ValueError(f"Unsupported input file type: {ext}")


def ensure_standard_logic_columns(df):
    """
    Create standard columns if the source file uses slightly different names
    or if the relevant columns are only present by position.
    """
    out = df.copy()
    raw_cols = list(out.columns)

    if "Transaction Date" not in out.columns and len(raw_cols) >= 1:
        out["Transaction Date"] = out.iloc[:, 0]
    if "Description" not in out.columns and len(raw_cols) >= 2:
        out["Description"] = out.iloc[:, 1]
    if "Amount" not in out.columns and len(raw_cols) >= 3:
        out["Amount"] = out.iloc[:, 2]
    if "C.D.Falg" not in out.columns and len(raw_cols) >= 4:
        out["C.D.Falg"] = out.iloc[:, 3]
    if "Reference No" not in out.columns and len(raw_cols) >= 5:
        out["Reference No"] = out.iloc[:, 4]
    if "Value Date" not in out.columns and len(raw_cols) >= 6:
        out["Value Date"] = out.iloc[:, 5]
    if "Branch Name" not in out.columns and len(raw_cols) >= 7:
        out["Branch Name"] = out.iloc[:, 6]
    if "Running Balance" not in out.columns and len(raw_cols) >= 8:
        out["Running Balance"] = out.iloc[:, 7]

    return out


def clear_target_sheet(ws, start_row=2, end_col=DEST_SPLIT_REFUND_COL):
    if ws.max_row < start_row:
        return

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.value = None


def copy_raw_data_to_target(ws, df):
    if df is None or df.empty:
        return

    src_cols = list(df.columns[:RAW_COL_COUNT])

    for row_idx, (_, record) in enumerate(df.iterrows(), start=2):
        for col_idx in range(RAW_COL_COUNT):
            col_name = src_cols[col_idx]
            value = record[src_cols[col_idx]] if col_idx < len(src_cols) else None
            #ye old vale logic ke hisab se tha sirf 1 line 
            #ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            col_name_norm = normalize_text(col_name)
            
            
            





         
            
            # 🔥 FIX: handle Amount / Transaction Amount with precision
            # if "AMOUNT" in col_name_norm:
            #     try:
            #         cell.value = Decimal(str(value))
            #         #cell.value = float(Decimal(str(value)))
            #     except:
            #         cell.value = value
            #     cell.number_format = '#,##0.00'
            # else:
            #     cell.value = value
            if "AMOUNT" in col_name_norm:
                try:
                    val = Decimal(str(value))
                    cell.value = float(val)
                    cell.number_format = 'General'   # 🔥 force no formatting
                    cell._style = None              # 🔥 reset template style
                        
                except:
                    cell.value = value
                    cell.number_format = 'General'
            else:
                cell.value = value
                        
                        
                
                
                
                
            
            # if "AMOUNT" in col_name_norm:
            #     try:
            #         val = Decimal(str(value))
                    
            #         cell.value = float(val)
            #         cell.number_format = '0.00'
            #         cell._style = None   # 🔥 reset template formatting
            #     except:
            #             cell.value = value
            #             cell.number_format = '0.00'
            # else:
            #                 cell.value = value
                            

            






            
            




def apply_tag_logic(df, refund_rrn_map=None):
    """
    Order:
      MPR (C only) -> Chargeback (D only) -> Payout (D only) ->
      MDR (D only) -> FD (C only) -> Knock Off (override) -> Refund (D only)
      Extra SP mapping fills blank SP only.
      Refund rows can additionally be enriched from the HDFC UPI refund file.
    """
    if "Tranaction Tag" not in df.columns:
        df["Tranaction Tag"] = ""
    if "SP Identifier/MID" not in df.columns:
        df["SP Identifier/MID"] = ""
    if "Split Refunds" not in df.columns:
        df["Split Refunds"] = ""

    for idx in df.index:
        description = df.at[idx, "Description"] if "Description" in df.columns else ""
        flag = df.at[idx, "C.D.Falg"] if "C.D.Falg" in df.columns else ""

        tag = None
        sp = None

         #1st ye run hoga 
        if row_tag_allowed(flag, "MPR"):
            tag, sp = get_mpr_credit_from_sp(description)
            if tag is None:
                tag, sp = get_fd_mapping(description)


        #2nd ye run hoga 
        if tag is None and row_tag_allowed(flag, "CHARGEBACK"):
            tag, sp = get_chargeback(description)
 
    
        #3rd ye run hoga 
        if tag is None and row_tag_allowed(flag, "MDR"):
            tag, sp = get_mdr(description)

        #4th ye run hoga 
        if tag is None and row_tag_allowed(flag, "PAYOUT"):
            tag, sp = get_payout(description)



        # if tag is None and row_tag_allowed(flag, "MDR"):
        #     tag, sp = get_mdr(description)


        #end mai ye run hoga
        if tag is None and row_tag_allowed(flag, "REFUND"):
            tag, sp = get_refund(description, current_tag="")

        if tag is not None:
            df.at[idx, "Tranaction Tag"] = tag
            df.at[idx, "SP Identifier/MID"] = sp if sp is not None else ""



            # 🔥 REFUND OVERRIDE (for 1CREDIT VOUCHER etc.)
    for idx in df.index:
        description = df.at[idx, "Description"]
        
        refund_tag, refund_sp = get_refund(
            description,
            df.at[idx, "Tranaction Tag"],
            df.at[idx, "SP Identifier/MID"]
            )
        if refund_tag == "Refund":
            df.at[idx, "Tranaction Tag"] = "Refund"
            df.at[idx, "SP Identifier/MID"] = ""
            


    df = mark_knock_off_rows(df)

    # New requirement: for rows already tagged as Refund, fill SP Identifier/MID
    # from the HDFC UPI refund file only when the SP cell is blank.
    if refund_rrn_map and "Reference No" in df.columns:
        for idx in df.index:
            current_tag = safe_str(df.at[idx, "Tranaction Tag"]).strip()
            current_sp = safe_str(df.at[idx, "SP Identifier/MID"]).strip()

            if current_tag == "Refund" and not current_sp:
                ref_no = df.at[idx, "Reference No"]
                resolved_sp = resolve_refund_sp_from_rrn(ref_no, refund_rrn_map)
                if resolved_sp:
                    df.at[idx, "SP Identifier/MID"] = resolved_sp

    if "Description" in df.columns:
        for idx in df.index:
            description = df.at[idx, "Description"]
            mapped_sp = get_sp_identifier_mid_mapping(description)
            if mapped_sp:
                current_sp = safe_str(df.at[idx, "SP Identifier/MID"]).strip()
                if not current_sp:
                    df.at[idx, "SP Identifier/MID"] = mapped_sp

    return df


def write_logic_columns_to_sheet(ws, df):
    for row_idx, (_, record) in enumerate(df.iterrows(), start=2):
        ws.cell(row=row_idx, column=DEST_TAG_COL, value=record.get("Tranaction Tag", ""))
        ws.cell(row=row_idx, column=DEST_SP_COL, value=record.get("SP Identifier/MID", ""))
        ws.cell(row=row_idx, column=DEST_SPLIT_REFUND_COL, value=record.get("Split Refunds", ""))


def build_processed_workbook_and_df(input_file_path, template_path, refund_file_path=None, refund_rrn_map=None):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    if not os.path.exists(input_file_path):
        raise FileNotFoundError(f"Input file not found: {input_file_path}")

    df = read_input_file(input_file_path)
    if df is None or df.empty:
        raise ValueError("Input file is empty or could not be read.")

    df = ensure_standard_logic_columns(df)

    for col in ["Tranaction Tag", "SP Identifier/MID", "Split Refunds"]:
        if col not in df.columns:
            df[col] = ""

    if refund_rrn_map is None:
        refund_rrn_map = {}
        if refund_file_path:
            if not os.path.exists(refund_file_path):
                raise FileNotFoundError(f"Refund file not found: {refund_file_path}")
            refund_df = read_input_file(refund_file_path)
            refund_rrn_map = build_refund_rrn_map(refund_df)

    df = apply_tag_logic(df, refund_rrn_map=refund_rrn_map)

    wb = openpyxl.load_workbook(template_path)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Target sheet not found in template: {TARGET_SHEET_NAME}")

    ws = wb[TARGET_SHEET_NAME]
    clear_target_sheet(ws, start_row=2, end_col=DEST_SPLIT_REFUND_COL)
    copy_raw_data_to_target(ws, df)
    write_logic_columns_to_sheet(ws, df)

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    return wb, df


def process_workbook(input_file_path, template_path, output_file_path, refund_file_path=None, refund_rrn_map=None):
    wb, _ = build_processed_workbook_and_df(input_file_path, template_path, refund_file_path=refund_file_path, refund_rrn_map=refund_rrn_map)
    wb.save(output_file_path)
    return output_file_path


def process_workbook_to_bytes(input_file_path, template_path, refund_file_path=None, refund_rrn_map=None):
    wb, df = build_processed_workbook_and_df(input_file_path, template_path, refund_file_path=refund_file_path, refund_rrn_map=refund_rrn_map)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, df


def default_output_path(input_file_path):
    base, _ = os.path.splitext(input_file_path)
    return base + "_processed.xlsx"


# ---------------------------------------------------------------------
# CLI / Spyder entry point
# ---------------------------------------------------------------------
def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE_PATH
    template_path = sys.argv[2] if len(sys.argv) > 2 else TEMPLATE_PATH
    output_path = sys.argv[3] if len(sys.argv) > 3 else OUTPUT_FILE_PATH

    if not input_path:
        raise ValueError(
            "INPUT_FILE_PATH is empty. Set it in the script or pass input/template/output paths via command line."
        )

    if not output_path:
        output_path = default_output_path(input_path)

    saved_path = process_workbook(
        input_file_path=input_path,
        template_path=template_path,
        output_file_path=output_path,
    )
    print(f"Done. Output saved to: {saved_path}")


# ---------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------
def run_streamlit_app():
    import streamlit as st

    st.title("HDFC ESCROW Automation")

    input_uploads = st.file_uploader(
        "Upload Statement Excel Files",
        type=["xlsx", "xlsm", "xls", "csv"],
        accept_multiple_files=True,
    )
    template_upload = st.file_uploader("Upload Raw Excel File", type=["xlsx", "xlsm"])
    refund_uploads = st.file_uploader(
        "Upload HDFC UPI Refund Files",
        type=["xlsx", "xlsm", "xls", "csv"],
        accept_multiple_files=True,
    )

    if input_uploads:
        st.success(f"{len(input_uploads)} input file(s) uploaded successfully")

    if template_upload is not None:
        st.success("Raw file uploaded successfully")

    if refund_uploads:
        st.info(f"{len(refund_uploads)} HDFC UPI refund file(s) uploaded successfully")

    if input_uploads and template_upload is not None:
        if st.button("Process Files"):
            template_suffix = os.path.splitext(template_upload.name)[1] or ".xlsx"
            temp_template_path = None
            temp_refund_paths = []
            combined_refund_map = {}

            try:
                with st.spinner("Preparing files..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix=template_suffix) as tmp_tpl:
                        tmp_tpl.write(template_upload.getbuffer())
                        temp_template_path = tmp_tpl.name

                    if refund_uploads:
                        for refund_upload in refund_uploads:
                            refund_suffix = os.path.splitext(refund_upload.name)[1] or ".xlsx"
                            with tempfile.NamedTemporaryFile(delete=False, suffix=refund_suffix) as tmp_ref:
                                tmp_ref.write(refund_upload.getbuffer())
                                temp_refund_paths.append(tmp_ref.name)

                        combined_refund_map = build_refund_rrn_map_from_paths(temp_refund_paths)

                for input_upload in input_uploads:
                    input_suffix = os.path.splitext(input_upload.name)[1] or ".xlsx"
                    temp_input_path = None

                    try:
                        with st.spinner(f"Processing {input_upload.name}..."):
                            with tempfile.NamedTemporaryFile(delete=False, suffix=input_suffix) as tmp_in:
                                tmp_in.write(input_upload.getbuffer())
                                temp_input_path = tmp_in.name

                            output_bytes, processed_df = process_workbook_to_bytes(
                                temp_input_path,
                                temp_template_path,
                                refund_rrn_map=combined_refund_map,
                            )

                            st.success(f"Processing Done: {input_upload.name} ✅")

                            blank_df = processed_df[
                                processed_df["Tranaction Tag"].astype(str).str.strip().eq("")
                            ] if "Tranaction Tag" in processed_df.columns else pd.DataFrame()

                            st.subheader(f"Rows with blank Transaction Tag - {input_upload.name}")
                            if blank_df.empty:
                                st.info("No blank Transaction Tag rows found.")
                            else:
                                st.dataframe(blank_df)

                            download_name = os.path.splitext(input_upload.name)[0] + "_processed.xlsx"
                            st.download_button(
                                label=f"Download {input_upload.name}",
                                data=output_bytes,
                                file_name=download_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{input_upload.name}",
                            )
                    finally:
                        if temp_input_path and os.path.exists(temp_input_path):
                            try:
                                os.remove(temp_input_path)
                            except Exception:
                                pass
            finally:
                if temp_template_path and os.path.exists(temp_template_path):
                    try:
                        os.remove(temp_template_path)
                    except Exception:
                        pass
                for p in temp_refund_paths:
                    if p and os.path.exists(p):
                        try:
                            os.remove(p)
                        except Exception:
                            pass



if __name__ == "__main__":
    if "streamlit" in sys.modules:
        run_streamlit_app()
    else:
        main()
